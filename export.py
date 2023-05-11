
import os

from sqlite3 import connect
import names
from openpyxl import Workbook, load_workbook

from registery import Regdata
from setting import Dbset


class Generator(Regdata):
    """
    generate data based on software settings
    export data as file 
    display data on software window.
    """
    
    def __init__(self, duration:int =None) -> None:
        self.fullnames = [names.get_full_name() for i in range(1000)]
        self.fullname = names.get_full_name()
        self.usernames = self.fullname.split(" ")
        self.data = None
        self.__db = Dbset()
        self.period = duration
        if duration is None:
            self.period = 2
        super().__init__(self.usernames[1], mail_service_provider="gmail.com")
        self.usr = self.user()  
    
    def _recorddata(self):
        
        query = """insert into emails(fullname, username, address, password, active, registered) values(?, ?, ?, ?, ?, ?)"""
        params = (
            self.data["fullname"], self.data["username"], self.data["address"], self.data["password"], self.data["active"], self.data["registered"]
        )
        newcon = connect("datahouse.db")
        cur = newcon.cursor()
        cur.execute(query, params)
        newcon.commit()
        return
        
    def _fetchdata(self, offset=0, limit=500):

        self.__db._con.close()
        newcon = connect("datahouse.db")
        cur = newcon.cursor()
        cur.execute(f"SELECT * FROM emails limit {limit} offset {offset}")
        data = cur.fetchall()
        newcon.close()
        #print(dict(zip(data[0].keys(), data[0])))
        
        return data
        
    def filedata(self, limit=500):
        """
        file the data based on the desired file type
        export to desired location: local or remote
        """
        
        user_data = self._fetchdata(limit=limit)
        
        if not os.path.exists(path="emails.xlsx"):
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Generated Email address"
            sheet.cell(row=1, column=1).value = "Full Name"
            sheet.cell(row=1, column=2).value = "User Name"
            sheet.cell(row=1, column=3).value = "Email address"
            sheet.cell(row=1, column=4).value = "Password"
            sheet.cell(row=1, column=5).value = "Active Email"
            sheet.cell(row=1, column=6).value = "Registered"
            sheet.cell(row=1, column=7).value = "Date"
            new_row = 1
        else:
            wb = load_workbook("emails.xlsx")
            sheet = wb[0]
            new_row = sheet.maxmum_row()
        sheet.cell(row=new_row, column=1).value = user_data.get("fullname")
        sheet.cell(row=new_row, column=2).value = user_data.get("username")
        sheet.cell(row=new_row, column=3).value = user_data.get("address")
        sheet.cell(row=new_row, column=4).value = user_data.get("password")
        sheet.cell(row=new_row, column=5).value = user_data.get("active")
        sheet.cell(row=new_row, column=6).value = user_data.get("registered")
        sheet.cell(row=new_row, column=7).value = user_data.get("dated") 
        
    def displaydata(self):
        """
        prepare and display data on the window
        """
        data = tuple(self.data.values())
        return data
    
    def generate_user_data(self, retry:bool=False):
        """
        check for username availability
        user_data: dict organize user data in dict
        """
        
        user_data = {'fullname': self.fullname}
        
        if retry:
            self.name = self.usernames[0]
        user_data.update(self.usr)
        user_data["dated"] = "Live"
        self.data = user_data
        return user_data
    
    def generator(self):
        for fullname in self.fullnames:
            yield self.generate_user_data(fullname=fullname)
    
    def __call__(self, rand:int=None,):
        """Retry data generation with appending random number
        """
        user_data = {'fullname': self.fullname}
        usr = super().__call__(rand=rand)
        user_data.update(usr)
        user_data["dated"] = "Live"
        self.data= user_data
        return user_data
   
if __name__ == "__main__":
    gen = Generator()
    print(gen.generate_user_data())
    genr = gen()
    print(genr)
